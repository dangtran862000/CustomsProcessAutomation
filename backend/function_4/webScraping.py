from selenium import webdriver
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import time
import sys
import openpyxl


# Opening JSON file
with open('./data.json')  as f:
  data = json.load(f)

username = data['username']
password = data['password']
tax_code = data['taxcode']

appState = {
    "recentDestinations": [{"id": "Save as PDF", "origin": "local", "account": ""}],
    "selectedDestinationId": "Save as PDF",
    "version": 2,
}

profile = {
    "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}],
    "printing.print_preview_sticky_settings.appState": json.dumps(appState),
}


def clickAble(driver, select):
    try:
        element = WebDriverWait(driver, 10).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, select))
        )
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, select))
        )
        element.click()
    except:
        try:
            element = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, select))
            )
            element.click()
        except Exception as e:
            print("Failed process", select)
            return False
    return True


def checkExist(driver, select):
    try:
        element = driver.find_element(By.CSS_SELECTOR, select)
        return element
    except NoSuchElementException:
        return None


def getIdentifierNumber(driver, select):
    try:
        WebDriverWait(driver, 10).until(
            lambda driver: len(driver.find_element(By.CSS_SELECTOR, select).text) > 0
        )
    except TimeoutException:
        pass


def checkLogin(driver):
    # username
    driver.find_element(By.CSS_SELECTOR, "#pt1\\3A it1\\3A \\3A content").send_keys(
        username
    )

    # password
    driver.find_element(By.CSS_SELECTOR, "#pt1\\3A it2\\3A \\3A content").send_keys(
        password
    )

    # handle captcha
    # retry = 0
    # flag = False
    # for i in range(3):
    #     captcha = driver.find_element(By.CSS_SELECTOR, "#pt1\\3A it42\\3A \\3A content")
    #     captcha.clear()
    #     captcha_text = SaveImage(browser=driver)
    #     if len(captcha_text) != 3:
    #         pass
    #     captcha.send_keys(captcha_text)
    #     # Click "lay thong tin"
    #     clickAble("#pt1\\3A cbLogin a")
    #     time.sleep(4)

    #     element = checkExist(
    #         "#d1_msgDlg\:\:_cnt > div > table > tbody > tr > td > table > tbody > tr > td.x1e3 > div"
    #     )
    #     if element is not None:
    #         # Failed for Login
    #         if element.text == "Đăng nhập thất bại.":
    #             print("Failed to login")
    #             return False
    #         # Failed for Captcha
    #         if element.text == "Nhập sai mã xác nhận":
    #             # if error message appear
    #             clickAble("#d1_msgDlg_cancel a")
    #             # retry += 1
    #             # print("Attempt:", retry)
    #     else:
    #         flag = True
    #         break
    # manually input captcha
    # if flag == False:
    driver.find_element(By.CSS_SELECTOR, "#pt1\\3A it42\\3A \\3A content").clear()
    WebDriverWait(driver, 300).until(
        lambda driver: len(
            driver.find_element(
                By.CSS_SELECTOR, "#pt1\\3A it42\\3A \\3A content"
            ).get_attribute("value")
        )
        == 3
    )
    # Click "lay thong tin"
    clickAble(driver,"#pt1\\3A cbLogin a")

    try:
        element = checkExist(driver,
            "#d1_msgDlg\:\:_cnt > div > table > tbody > tr > td > table > tbody > tr > td.x1e3 > div"
        )
        if element.text == "Đăng nhập thất bại.":
            print("Failed to login")
            return False
        # Failed for Captcha
        if element.text == "Nhập sai mã xác nhận":
            print("Failed to handle captcha")
            return False
    except Exception as e:
        return True

    return True


def downloadFile(directory, numberOfCodes, options, targets):
    chrome_options = webdriver.EdgeOptions()
    chrome_options.add_experimental_option("prefs", profile)
    chrome_options.add_argument("--kiosk-printing")
    chrome_options.add_argument("--start-maximized")

    driver = webdriver.Edge(options=chrome_options)
    # driver.get('edge://settings/clearBrowserData')
    # driver.find_element("xpath",'//settings-ui').send_keys(Keys.ENTER)

    time.sleep(5)
    url_1 = "https://pus.customs.gov.vn/faces/Login"
    driver.get(url_1)
    # login process
    isLogin = checkLogin(driver)
    if isLogin == False:
        return False

    time.sleep(2)

    # Click "Dinh danh hang hoa"
    clickAble(driver, "#pt1\\3A dc7\\3A dinhDanhhangHoa a")

    # Click "Dinh danh hang hoa"
    clickAble(driver, "#pt1\\3A b1 a") == False

    time.sleep(5)

    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "#pt1\\3A soc1\\3A \\3A content"))
    )

    select = Select(
        driver.find_element(By.CSS_SELECTOR, "#pt1\\3A soc1\\3A \\3A content")
    )
    
    select.select_by_value("0")

    if targets == "target01":
        select.select_by_value("0")
    elif targets == "target02":
        select.select_by_value("1")
    else:
        select.select_by_value("2")

    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "#pt1\\3A soc2\\3A \\3A content"))
    )

    select = Select(
        driver.find_element(By.CSS_SELECTOR, "#pt1\\3A soc2\\3A \\3A content")
    )

    if options == "opt1":
        select.select_by_value("0")
    else:
        select.select_by_value("1")

    driver.find_element(By.CSS_SELECTOR, "#pt1\\3A it5\\3A \\3A content").send_keys(
        tax_code
    )

    clickAble(driver, "#pt1\\3A cb3 a")

    clickAble(driver, "#pt1\\3A b4 a")

    text = ""
    list = []

    try:
        text = driver.find_element(
            By.CSS_SELECTOR, "#pt1\\3A pc1\\3A t2\\3A 0\\3A c2 span"
        ).text
    except NoSuchElementException:
        pass
    
    # Open the Excel workbook for writing
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    if text != "":
    # Write the value of text to a new row in the first column
        sheet.cell(row=1, column=1, value=text)
        list.append(text)

    for i in range(int(numberOfCodes) - 1):
        text = ""
        clickAble(driver, "#pt1\\3A b1 a")
        clickAble(driver, "#pt1\\3A cb3 a")
        clickAble(driver, "#pt1\\3A b4 a")
        getIdentifierNumber(driver, "#pt1\\3A pc1\\3A t2\\3A 0\\3A c2 span")
        try:
            text = driver.find_element(By.CSS_SELECTOR, "#pt1\\3A pc1\\3A t2\\3A 0\\3A c2 span").text
        except NoSuchElementException:
            pass
        if text != "" and text != list[-1]:
            # Write the value of text to a new row in the first column
            sheet.cell(row=i + 2, column=1, value=text)
            list.append(text)

    if len(list) == int(numberOfCodes):
        print("success")
    else:
        print("failure")
    print(len(list))
    # with open(directory+"myfile.txt", "a+") as f:
    #     f.write(" ".join(list))

    workbook.save(directory + "UCR.xlsx")

    driver.quit()


def start(args):
    CDS_Folder = args[0]
    directory = f"{CDS_Folder}"
    numberOfCodes = args[1]
    options = args[2]
    targets = args[3]
    downloadFile(directory, numberOfCodes, options, targets)
    sys.stdout.flush()
