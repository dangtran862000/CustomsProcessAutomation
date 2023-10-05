import sys
from win32com import client
import re
import openpyxl
import os
import win32com.client as win32
import tkinter as tk
import PyPDF2
from PyPDF2 import PdfMerger
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import win32com.client as win32
from PyPDF2 import PdfWriter, PdfReader
import xlrd

to = sys.argv[1]
filePath =  sys.argv[2]
shippingdocsPath = sys.argv[3]
BillPath = sys.argv[4]


def convertListToString(list):
    POnum = ""

    for i in range(0, len(list)):
    
        if i == (len(list)-1):
            POnum += str(list[i])
        else:
            POnum += str(list[i]) + "; "

    # print(POnum)
    return POnum


def exportExceltoPDF(dir, typeFileName, fileNameInput, fileNameOutput):

    pathIn = dir +"/"+ fileNameInput
    pathOut = dir +"/"+ fileNameOutput
    # print(pathIn)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    
    excel.Interactive = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False

    if typeFileName==".xlsx":
        sheets = excel.Workbooks.Open(pathIn+".xlsx")
        work_sheets = sheets.Worksheets[1]
    else:
        sheets = excel.Workbooks.Open(pathIn+".xls")
        work_sheets = sheets.Worksheets[1]
     # Read Excel File
    
    # Convert into PDF File.
    filepath = pathOut+".pdf"
    if os.path.isfile(filepath):
        # print('Đã có file Thanh Cong')
        os.remove(filepath)
        work_sheets.ExportAsFixedFormat(0, pathOut)
        # print(pathOut)

    else:
        work_sheets.ExportAsFixedFormat(0, pathOut)
        # print(pathOut)

    sheets.Close(False)
    excel.Application.Quit()


def deleteBlankPage(inputFilePath, type):
    if type == "CL":
        pages_to_keep = [0] # page numbering starts from 0
        infile = PdfReader(inputFilePath, 'rb')
        output = PdfWriter()

        for i in pages_to_keep:
            p = infile.pages[i] 
            output.add_page(p)

        with open(inputFilePath, 'wb') as f:
            output.write(f)
    else:
        file1 = open(inputFilePath, 'rb')
        ReadPDF = PyPDF2.PdfReader(file1)
        #No of pages initially
        pages = len(ReadPDF.pages)
        # print(pages)

        #Creating new file which do not conatin any empty pages
        output = PyPDF2.PdfWriter()

        for i in range(pages):
            ReadPDF = PyPDF2.PdfReader(file1)
            pageObj = ReadPDF.pages[i]
            text = pageObj.extract_text()
            
            if (len(text) > 0):
                output. add_page(pageObj)
                
        with open(inputFilePath, 'wb') as f:
                output.write(f)
        output.close()
        f.close()
        file1.close()


def findShippingDoc(Invoice_no, CDS_no):
    res = os.listdir(dir_shippingdocs_currently)
    listFileShippingDocName = []
    for i in res:
        fileShippingDoc = os.path.basename(i)
        listFileShippingDocName.append(fileShippingDoc)
    # print(listFileShippingDocName)

    fileNameCheck = str(Invoice_no)
    # print(fileNameCheck)

    for i in listFileShippingDocName:
        i_split = str(os.path.splitext(i)[0]).split('_Offical' or '_offical')[0]
        i_split = i_split.split("_")[1]
        if (i_split == fileNameCheck):
            # print(i_split)
            fileShippingDoc_path = dir_shippingdocs_currently+ "/" + i
            return str(fileShippingDoc_path)
        else:
            pass
    return None


def findBill(shippingDocPath):
    fileShippingDocName = os.path.basename(shippingDocPath)
    # print(str(os.path.splitext(fileShippingDocName)[0]).split("_")[0])
    res = os.listdir(dir_bill_currently)
    listFileBill = []
    for i in res:
        fileBill = os.path.basename(i)
        listFileBill.append(fileBill)
    # print(listFileShippingDocName)

    fileNameCheck = str(os.path.splitext(fileShippingDocName)[0]).split("_")[0]
    # print(fileNameCheck)

    for i in listFileBill:
        i_split = str(os.path.splitext(i)[0])
        i_split = i_split.split("_")[0]
        if (i_split == fileNameCheck):
            # print(i_split)
            fileBill_path = dir_bill_currently+ "/" + i
            # print(fileBill_path)
            return str(fileBill_path)
        else:
            pass
    return None

def mergedPDF(pathCheckList, pathCDS, pathShippingDoc, pathBill, CDS_no, ModeCode, YYYY):
    merger = PdfMerger()
    result1 = pathCheckList
    result2 = pathCDS
    result3 = pathShippingDoc
    result4 = pathBill

    # print(result1)
    # print(result2)
    # print(result3)
     
    pdf_files = [str(result1), str(result2), str(result4), str(result3)]

    #Iterate over the list of the file paths
    for pdf_file in pdf_files:
        #Append PDF files
        merger.append(pdf_file)

    #Write out the merged PDF file
    merger.write(completed_dir+"/"+CDS_no+"_"+ModeCode+"_"+YYYY+".pdf")
    merger.close()
    return completed_dir+"/"+CDS_no+"_"+ModeCode+"_"+YYYY+".pdf"


def createCDS_package(NameSign):
    
    # global messageCDS_no, messageDate, messageInvoice_no, messageMLH_no, messageWB_no, messagePO_No, messagePIC_Checklist, messageStatus, buttonViewFile, clearall
    global finalCDSPackage_path, completed_dir
    if NameSign == "":
        pass
    else:    
        # import pyexcel as p


        CDS_dest_path=dir_currently+'/'+os.path.splitext(CDSFileName)[0]+'.xlsx'

        # print(os.path.splitext(CDSFileName)[1])

        if os.path.splitext(CDSFileName)[1] == ".xls":
            # p.save_book_as(file_name=dir_currently+'/'+CDSFileName,
            #     dest_file_name=CDS_dest_path)
            CDS_dest_path=dir_currently+'/'+os.path.splitext(CDSFileName)[0]+'.xls'

        # Give the location of the file
        path_CheckList = os.getcwd()+"/backend/function_2/document/export/"+"/Check-list-template_Export.xlsx"
        # print(path_CDS)
        # To open the workbook
        # workbook object is created
                

        try:
            wb_obj_open = xlrd.open_workbook(CDS_dest_path)
        except:
            print("error")
            return ""
        
        sheet_obj = wb_obj_open.sheet_by_index(0)
        
        wb_obj_write =  openpyxl.load_workbook(path_CheckList)
        sheet_obj_write = wb_obj_write.active
        try:
            cell_CDS_No = sheet_obj.cell(rowx = 3, colx = 4)
            cell_MLH_No = sheet_obj.cell(rowx = 5, colx = 11)
            cell_DATE = sheet_obj.cell(rowx = 7, colx = 5)
            cell_WB_No = sheet_obj.cell(rowx = 38, colx = 7)
            cell_Invoice_No = sheet_obj.cell(rowx = 48, colx = 17)
            
            
        except:
            print("CDS File name '"+ CDSFileName +"' invalid")
            print(CDSFileName)
            return ""
        
        

        if (cell_CDS_No.value == "") or (cell_MLH_No == "") or (cell_WB_No == "") or (cell_Invoice_No == ""):
            print("CDS File name '"+ CDSFileName +"' invalid")
            print(CDSFileName)
            return ""



        if (cell_CDS_No.value == None) or (cell_MLH_No == None) or (cell_WB_No == None) or (cell_Invoice_No == None):
            print("CDS File name '"+ CDSFileName +"' invalid")
            print(CDSFileName)
            return ""
        

        # Print value of cell object
        # using the value attribute
        
        
        if findShippingDoc(Invoice_no(cell_Invoice_No),cell_CDS_No.value) == None:
            print("CDS File name '"+ CDSFileName +"' can not find the shipping docs relevant")
            print(CDSFileName)
            return ""
        if findBill(findShippingDoc(Invoice_no(cell_Invoice_No),cell_CDS_No.value)) == None:
            print("CDS File name '"+ CDSFileName +"' can not find the bill relevant")
            print(CDSFileName)
            return ""

        else:
            
            try:
                cell_ISM_DN_CIC_CL = sheet_obj_write.cell(row = 2, column = 3)
                cell_PO_ASM_CDS_CL = sheet_obj_write.cell(row = 3, column = 3)
                cell_NameSign1_CL = sheet_obj_write.cell(row = 25, column = 1)
                
                stringPO_ASM_CDS_CL = str(int(cell_CDS_No.value))+"_"+str(MLH(cell_MLH_No))+"_"+str(date(cell_DATE))
                
                cell_PO_ASM_CDS_CL.value = stringPO_ASM_CDS_CL
                cell_ISM_DN_CIC_CL.value = Invoice_no(cell_Invoice_No)
                
            except:
                print("CDS File name '"+ CDSFileName +"' invalid")
                print(CDSFileName)
                return ""
            cell_NameSign1_CL.value = "Name & Signature: " + NameSign
       

            # messageCDS_no = Label (frame, text="\n\nCDS number: " +str(cell_CDS_No.value))
            # messageCDS_no.pack()
            # messageMLH_no = Label (frame, text="Mode code (MLH): "+MLH(cell_MLH_No))
            # messageMLH_no.pack()
            # messageWB_no = Label (frame, text="Waybill number: "+WB(cell_WB_No))
            # messageWB_no.pack()
            # messageInvoice_no = Label (frame, text="Invoice number: "+Invoice_no(cell_Invoice_No))
            # messageInvoice_no.pack()
            # messagePO_No = Label (frame, text="PO number: "+convertListToString(cell_PO_No_list))
            # messagePO_No.pack()
            # messageDate = Label (frame, text="Date CDS'create: "+date(cell_DATE))
            # messageDate.pack()
            # messagePIC_Checklist = Label (frame, text="PIC name: "+NameSign)
            # messagePIC_Checklist.pack()

            fileName =str(WB(cell_WB_No))+"_"+str(Invoice_no(cell_Invoice_No))+"_"+str(cell_CDS_No.value)
            pathCL = dir_currently+"/"+fileName
            
            # Open Microsoft Excel
            CDSBaseFileName = os.path.splitext(CDSFileName)[0]
            # print(CDSBaseFileName)

            completed_dir = dir_currently + "/Completed CDS package"
            checkExist = completed_dir+"/"+str(int(cell_CDS_No.value))+"_"+str(MLH(cell_MLH_No))+"_"+str(date(cell_DATE))+".pdf"
            if os.path.exists(completed_dir):
                pass
            else:
                os.mkdir(completed_dir)
            
            if os.path.exists(checkExist):
                print("Tao CDS package thanh cong !!!")
                return ""
            else:
                wb_obj_write.save(pathCL+".xlsx")
                wb_obj_write.close()
                try:
                    exportExceltoPDF(dir_currently,".xlsx",fileName,fileName+"_CL")
                except:
                    print("CDS File name '"+ CDSFileName +"' can not create package !!! Please close the Microsoft Excel before create paackage !!!")
                    print(CDSFileName)
                    os.remove(pathCL+".xlsx")
                    return ""

                if os.path.splitext(CDSFileName)[1] == ".xls":
                    exportExceltoPDF(dir_currently, ".xls",CDSBaseFileName,fileName)
                else:
                    exportExceltoPDF(dir_currently, ".xlsx",CDSBaseFileName,fileName)
                

                deleteBlankPage(pathCL+"_CL"+".pdf","CL")
                deleteBlankPage(pathCL+".pdf","CDS")

                
                
                finalCDSPackage_path = mergedPDF(pathCL+"_CL"+".pdf",pathCL+".pdf",findShippingDoc(Invoice_no(cell_Invoice_No),cell_CDS_No.value), findBill(findShippingDoc(Invoice_no(cell_Invoice_No),cell_CDS_No.value)),
                                                  str(int(cell_CDS_No.value)), str(MLH(cell_MLH_No)), str(date(cell_DATE)))
                
                # findBill(findShippingDoc(Invoice_no(cell_Invoice_No),cell_CDS_No.value))

                if os.path.isfile(pathCL+".xlsx"):
                    os.remove(pathCL+".xlsx")
                    os.remove(pathCL+".pdf")
                    os.remove(pathCL+"_CL"+".pdf")
                    print('Tao CDS package thanh cong !!!')

                def startFile():
                    os.startfile(finalCDSPackage_path)

                # messageStatus = Label (frame, text="\n\nSuscessfully create the CDS package file !", font='Helvetica 10 bold')
                # messageStatus.pack()
                # buttonViewFile = tk.Button(frame, text="View the CDS package file", command=startFile)
                # buttonViewFile.pack()
                # clearall = Button(frame, text='Create new package', command=resetALL)
                # clearall.pack()

                print(cell_CDS_No.value)
                print(MLH(cell_MLH_No))
                print(date(cell_DATE))
                print(WB(cell_WB_No))
                print(Invoice_no(cell_Invoice_No))
                print(NameSign)
                print(os.path.basename(finalCDSPackage_path))

                return str(cell_CDS_No.value), MLH(cell_MLH_No), WB(cell_WB_No), Invoice_no(cell_Invoice_No), date(cell_DATE), NameSign


def MLH(value):
    origin_value = value.value
    origin_value.split()
    MLH_code = origin_value.split()[0]
    return MLH_code


def date(value):
    return value.value.split()[0].split('/')[2]


def WB(value):
    return re.split(r'[`\-=~!@#$%^&*()_+\[\]{};\'\\:"|<,./<>?]', value.value)[0]


def Invoice_no(value):
    cell_Invoice_No = value.value
    # if(bool(re.search('/',cell_Invoice_No))==True):
    #     return re.sub('/','-',cell_Invoice_No)
    # else:
    return cell_Invoice_No


# def printInput():
#     inpName = inputtxt.get(1.0, "end-1c")
#     print(inpName)
#     return inpName


# def resetALL():
#     messageInvoice_no.destroy()
#     messageCDS_no.destroy() 
#     messageDate.destroy()
#     messageMLH_no.destroy()
#     messageWB_no.destroy()
#     messagePO_No.destroy()
#     messagePIC_Checklist.destroy()
#     messageStatus.destroy()
#     buttonViewFile.destroy()
#     messageFilePath.destroy()
#     clearall.destroy()


def browse_button(filename):
    global dir_currently,CDSFileName
    # Allow user to select a directory and store it in global var
    # called folder_path
    dir_currently = os.path.dirname(filename)
    # print(dir_currently)
    CDSFileName = os.path.basename(filename)
    # messageFilePath = Label (frame, text="\nCDS Package Folder: "+str(filename))
    # messageFilePath.pack()
    return dir_currently

def browse_shippingdocs_button(filename):
    global dir_shippingdocs_currently
    # Allow user to select a directory and store it in global var
    # called folder_path
    dir_shippingdocs_currently = os.path.dirname(filename)
    # print(dir_currently)
    # messageFilePath = Label (frame, text="\nCDS Package Folder: "+str(filename))
    # messageFilePath.pack()
    return dir_shippingdocs_currently

def browse_bill_button(filename):
    global dir_bill_currently
    # Allow user to select a directory and store it in global var
    # called folder_path
    dir_bill_currently = os.path.dirname(filename)
    # print(dir_currently)
    # messageFilePath = Label (frame, text="\nCDS Package Folder: "+str(filename))
    # messageFilePath.pack()
    return dir_bill_currently

def get_value(to, filePath):
    browse_button(filePath)
    browse_shippingdocs_button(shippingdocsPath)
    browse_bill_button(BillPath)
    MLH = createCDS_package(to)


get_value(to, filePath)
print(browse_shippingdocs_button(shippingdocsPath))
sys.stdout.flush()